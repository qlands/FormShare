from lxml import etree
import os
import glob
import io
import json
import mimetypes
import shutil
from hashlib import md5
from pyramid.httpexceptions import HTTPNotFound
from uuid import uuid4
from subprocess import Popen, PIPE, check_call, CalledProcessError
from pyramid.response import FileResponse
from formshare.processes.db import assistant_has_form, get_assistant_forms, get_project_id_from_name, add_new_form, \
    form_exists, get_form_directory, get_form_xml_file, get_form_xls_file, update_form, get_form_data, \
    get_form_schema, get_submission_data, add_submission, add_json_log, update_json_status, add_json_history, \
    form_file_exists, get_url_from_file, set_file_with_error, update_file_info, get_project_from_assistant, \
    get_form_files, get_project_code_from_id, get_form_geopoint, get_media_files, add_file_to_form, \
    add_submission_same_as
import uuid
import datetime
import re
import sys
from bs4 import BeautifulSoup
from sqlalchemy import exc
from decimal import Decimal
from zope.sqlalchemy import mark_changed
from pyxform import xls2xform
from openpyxl import load_workbook
from pyxform.errors import PyXFormError
from formshare.processes.storage import get_stream, response_stream, store_file, delete_stream
from pyramid.response import Response
from urllib.request import urlretrieve
from pyramid.httpexceptions import HTTPFound
from formshare.processes.elasticsearch.repository_index import create_dataset_index, add_dataset

import logging
log = logging.getLogger(__name__)

__all__ = ['generate_form_list', 'generate_manifest', 'get_form_list',
           'get_manifest', 'get_xml_form', 'get_media_file', 'get_submission_file', 'build_odata_service',
           'build_database', 'create_repository', 'move_media_files', 'convert_xml_to_json', 'store_submission',
           'get_html_from_diff', 'generate_diff', 'store_new_version', 'restore_from_revision', 'push_revision',
           'get_tables_from_form', 'get_fields_from_table', 'get_table_desc', 'is_field_key', 'get_request_data',
           'update_data', 'upload_odk_form', 'update_form_title', 'retrieve_form_file', 'cache_external_file',
           'get_odk_path', 'store_file_in_directory', 'update_odk_form']


def get_odk_path(request):
    repository_path = request.registry.settings['repository.path']
    if not os.path.exists(repository_path):
        os.makedirs(repository_path)
    return os.path.join(repository_path, *["odk"])


def cache_external_file(request, project, form, file_name):
    if form_file_exists(request, project, form, file_name):
        file_url, last_download = get_url_from_file(request, project, form, file_name)
        if file_url is not None:
            download = False
            if last_download is not None:
                lapsed = datetime.datetime.now() - last_download
                minutes_lapsed = divmod(lapsed.days * 86400 + lapsed.seconds, 60)
                minutes_lapsed = minutes_lapsed[0]
                if minutes_lapsed > 5:
                    download = True
            else:
                download = True
            if download:
                repository_path = request.registry.settings['repository.path']
                downloads_path = os.path.join(repository_path, *["downloads"])
                if not os.path.exists(downloads_path):
                    os.makedirs(downloads_path)
                target_file = os.path.join(downloads_path, *[str(uuid.uuid4()) + ".tmp"])
                try:
                    log.info("Retrieving file {}".format(file_url))
                    urlretrieve(file_url, target_file)
                except Exception as e:
                    set_file_with_error(request, project, form, file_name)
                    log.error(
                        "Error {} while downloading external file {} for form {}".format(str(e), file_name, form))
                    raise HTTPNotFound
                bucket_id = project + form
                bucket_id = md5(bucket_id.encode('utf-8')).hexdigest()
                file = open(target_file, 'rb')
                store_file(request, bucket_id, file_name, file)
                file.close()
                file = open(target_file, 'rb')
                md5sum = md5(file.read()).hexdigest()
                file.close()
                update_file_info(request, project, form, file_name, md5sum)
    else:
        raise HTTPNotFound


def store_file_in_directory(request, project, form, file_name, directory):
    if form_file_exists(request, project, form, file_name):
        file_url, last_download = get_url_from_file(request, project, form, file_name)
        if file_url is None:
            bucket_id = project + form
            bucket_id = md5(bucket_id.encode('utf-8')).hexdigest()
            stream = get_stream(request, bucket_id, file_name)
            if stream is not None:
                target_file = os.path.join(directory, *[file_name])
                file = open(target_file, 'wb')
                file.write(stream.read())
                file.close()
        else:
            download = False
            if last_download is not None:
                lapsed = datetime.datetime.now() - last_download
                minutes_lapsed = divmod(lapsed.days * 86400 + lapsed.seconds, 60)
                minutes_lapsed = minutes_lapsed[0]
                if minutes_lapsed > 5:
                    download = True
            else:
                download = True
            if download:
                repository_path = request.registry.settings['repository.path']
                downloads_path = os.path.join(repository_path, *["downloads"])
                if not os.path.exists(downloads_path):
                    os.makedirs(downloads_path)
                target_file = os.path.join(downloads_path, *[str(uuid.uuid4()) + ".tmp"])
                try:
                    log.info("Retrieving file {}".format(file_url))
                    urlretrieve(file_url, target_file)
                except Exception as e:
                    set_file_with_error(request, project, form, file_name)
                    log.error(
                        "Error {} while downloading external file {} for form {}".format(str(e), file_name, form))
                bucket_id = project + form
                bucket_id = md5(bucket_id.encode('utf-8')).hexdigest()
                file = open(target_file, 'rb')
                store_file(request, bucket_id, file_name, file)
                file.close()
                file = open(target_file, 'rb')
                md5sum = md5(file.read()).hexdigest()
                file.close()
                update_file_info(request, project, form, file_name, md5sum)

            bucket_id = project + form
            bucket_id = md5(bucket_id.encode('utf-8')).hexdigest()
            stream = get_stream(request, bucket_id, file_name)
            if stream is not None:
                target_file = os.path.join(directory, *[file_name])
                file = open(target_file, 'wb')
                file.write(stream.read())
                file.close()


def retrieve_form_file(request, project, form, file_name):
    if form_file_exists(request, project, form, file_name):
        file_url, last_download = get_url_from_file(request, project, form, file_name)
        if file_url is None:
            bucket_id = project + form
            bucket_id = md5(bucket_id.encode('utf-8')).hexdigest()
            stream = get_stream(request, bucket_id, file_name)
            if stream is not None:
                response = Response()
                return response_stream(stream, file_name, response)
            else:
                raise HTTPNotFound
        else:
            download = False
            if last_download is not None:
                lapsed = datetime.datetime.now() - last_download
                minutes_lapsed = divmod(lapsed.days * 86400 + lapsed.seconds, 60)
                minutes_lapsed = minutes_lapsed[0]
                if minutes_lapsed > 5:
                    download = True
            else:
                download = True
            if download:
                repository_path = request.registry.settings['repository.path']
                downloads_path = os.path.join(repository_path, *["downloads"])
                if not os.path.exists(downloads_path):
                    os.makedirs(downloads_path)
                target_file = os.path.join(downloads_path, *[str(uuid.uuid4()) + ".tmp"])
                try:
                    log.info("Retrieving file {}".format(file_url))
                    urlretrieve(file_url, target_file)
                except Exception as e:
                    set_file_with_error(request, project, form, file_name)
                    log.error(
                        "Error {} while downloading external file {} for form {}".format(str(e), file_name, form))
                    return HTTPFound(file_url)
                bucket_id = project + form
                bucket_id = md5(bucket_id.encode('utf-8')).hexdigest()
                file = open(target_file, 'rb')
                store_file(request, bucket_id, file_name, file)
                file.close()
                file = open(target_file, 'rb')
                md5sum = md5(file.read()).hexdigest()
                file.close()
                update_file_info(request, project, form, file_name, md5sum)

            bucket_id = project + form
            bucket_id = md5(bucket_id.encode('utf-8')).hexdigest()
            stream = get_stream(request, bucket_id, file_name)
            if stream is not None:
                response = Response()
                return response_stream(stream, file_name, response)
            else:
                raise HTTPNotFound
    else:
        raise HTTPNotFound


def update_form_title(request, project, form, title):
    xml_file = get_form_xml_file(request, project, form)
    xlsx_file = get_form_xls_file(request, project, form)
    tree = etree.parse(xml_file)
    root = tree.getroot()
    h_nsmap = root.nsmap['h']
    form_title = root.findall(".//{" + h_nsmap + "}title")
    if form_title:
        form_title[0].text = title
    tree.write(xml_file, pretty_print=True, xml_declaration=True, encoding="utf-8")

    form_directory = get_form_directory(request, project, form)
    md5(open(xml_file, 'rb').read()).hexdigest()
    odk_dir = get_odk_path(request)
    json_file = os.path.join(odk_dir, *["forms", form_directory, form.lower() + ".json"])
    with open(json_file, 'r') as f:
        json_metadata = json.load(f)
        json_metadata["name"] = title
        json_metadata['hash'] = 'md5:' + md5(open(xml_file, 'rb').read()).hexdigest()
        json_metadata["descriptionText"] = title
    with open(json_file, "w", ) as outfile:
        json_string = json.dumps(json_metadata, indent=4, ensure_ascii=False)
        outfile.write(json_string)

    wb = load_workbook(xlsx_file)
    sheet = wb['settings']
    for y in range(sheet.min_column, sheet.max_column+1):
        if sheet.cell(row=1, column=y).value == "form_title":
            sheet.cell(row=2, column=y, value=title)
    wb.save(xlsx_file)


def get_geopoint_variable_from_xlsx(xlsx_file):
    wb = load_workbook(xlsx_file)
    sheet = wb['survey']
    type_column = -1
    name_column = -1
    groups = []
    repeats = []
    for y in range(sheet.min_column, sheet.max_column+1):
        if sheet.cell(row=1, column=y).value == "type":
            type_column = y
        if sheet.cell(row=1, column=y).value == "name":
            name_column = y
    for x in range(2, sheet.max_row+1):
        variable_type = sheet.cell(x, type_column).value
        if variable_type is not None:
            variable_name = sheet.cell(x, name_column).value
            variable_type = " ".join(variable_type.split())
            variable_type = variable_type.lower()
            if variable_type == 'begin repeat':
                repeats.append(variable_name)
            if variable_type == 'end repeat':
                try:
                    del repeats[-1]
                except IndexError:
                    log.error("It seems that the file {} is closing an invalid group".format(xlsx_file))
            if variable_type == 'begin group':
                groups.append(variable_name)
            if variable_type == 'end group':
                try:
                    del groups[-1]
                except IndexError:
                    log.error("It seems that the file {} is closing an invalid group".format(xlsx_file))

            if variable_type == 'geopoint':
                if len(repeats) == 0:
                    if len(groups) > 0:
                        return "/".join(groups) + '/' + variable_name
                    else:
                        return variable_name
    return None


def upload_odk_form(request, project_id, user_id, odk_dir, form_data):
    uid = str(uuid.uuid4())
    form_directory = uid[-12:]
    paths = ['tmp', uid]
    os.makedirs(os.path.join(odk_dir, *paths))

    input_file = request.POST['xlsx'].file
    input_file_name = request.POST['xlsx'].filename

    paths = ['tmp', uid, input_file_name]
    file_name = os.path.join(odk_dir, *paths)

    input_file.seek(0)
    with open(file_name, 'wb') as permanent_file:
        shutil.copyfileobj(input_file, permanent_file)

    input_file.close()

    parts = os.path.splitext(input_file_name)

    paths = ['tmp', uid, parts[0]+'.xml']
    xml_file = os.path.join(odk_dir, *paths)

    try:
        xls2xform.xls2xform_convert(file_name, xml_file)

        # Check if the conversion created itemsets.csv
        output_dir = os.path.split(xml_file)[0]
        itemsets_csv = os.path.join(output_dir, "itemsets.csv")
        if not os.path.isfile(itemsets_csv):
            itemsets_csv = ""

        tree = etree.parse(xml_file)
        root = tree.getroot()
        nsmap = root.nsmap[None]
        h_nsmap = root.nsmap['h']
        eid = root.findall(".//{" + nsmap + "}" + parts[0])
        if eid:
            form_id = eid[0].get("id")
            if re.match(r'^[A-Za-z0-9_]+$', form_id):
                form_title = root.findall(".//{" + h_nsmap + "}title")
                if not form_exists(request, project_id, form_id):
                    paths = ['forms', form_directory, 'media']
                    if not os.path.exists(os.path.join(odk_dir, *paths)):
                        os.makedirs(os.path.join(odk_dir, *paths))

                        paths = ['forms', form_directory, 'submissions']
                        os.makedirs(os.path.join(odk_dir, *paths))

                        paths = ['forms', form_directory, 'submissions', 'logs']
                        os.makedirs(os.path.join(odk_dir, *paths))

                        paths = ['forms', form_directory, 'submissions', 'maps']
                        os.makedirs(os.path.join(odk_dir, *paths))

                        paths = ['forms', form_directory, 'repository']
                        os.makedirs(os.path.join(odk_dir, *paths))
                        paths = ['forms', form_directory, 'repository', 'temp']
                        os.makedirs(os.path.join(odk_dir, *paths))

                    xls_file_fame = os.path.basename(file_name)
                    xml_file_name = os.path.basename(xml_file)
                    paths = ['forms', form_directory, xls_file_fame]
                    final_xls = os.path.join(odk_dir, *paths)
                    paths = ['forms', form_directory, xml_file_name]
                    final_xml = os.path.join(odk_dir, *paths)
                    shutil.copyfile(file_name, final_xls)
                    shutil.copyfile(xml_file, final_xml)
                    geopoint = get_geopoint_variable_from_xlsx(final_xls)
                    form_data['project_id'] = project_id
                    form_data['form_id'] = form_id
                    form_data['form_name'] = form_title[0].text
                    form_data['form_cdate'] = datetime.datetime.now()
                    form_data['form_directory'] = form_directory
                    form_data['form_accsub'] = 1
                    form_data['form_testing'] = 1
                    form_data['form_xlsfile'] = final_xls
                    form_data['form_xmlfile'] = final_xml
                    form_data['form_pubby'] = user_id
                    form_data['form_geopoint'] = geopoint

                    added, message = add_new_form(request, form_data)
                    if not added:
                        return added, message

                    # If we have itemsets.csv add it as media files
                    if itemsets_csv != "":
                        with open(itemsets_csv, "rb", ) as itemset_file:
                            md5sum = md5(itemset_file.read()).hexdigest()
                            added, message = add_file_to_form(request, project_id, form_id, 'itemsets.csv', None, True,
                                                              md5sum)
                            if added:
                                itemset_file.seek(0)
                                bucket_id = project_id + form_id
                                bucket_id = md5(bucket_id.encode('utf-8')).hexdigest()
                                store_file(request, bucket_id, 'itemsets.csv', itemset_file)

                        #paths = ['forms', form_directory, 'media', 'itemsets.csv']
                        #itemset_file = os.path.join(odk_dir, *paths)
                        #shutil.copyfile(itemsets_csv, itemset_file)

                    paths = ['forms', form_directory, parts[0]+".json"]
                    json_file = os.path.join(odk_dir, *paths)

                    metadata = {"formID": form_id, "name": form_title[0].text, "majorMinorVersion": "",
                                "version": datetime.datetime.now().strftime("%Y%m%d"),
                                "hash": 'md5:' + md5(open(final_xml, 'rb').read()).hexdigest(),
                                "descriptionText": form_title[0].text}

                    with open(json_file, "w",) as outfile:
                        json_string = json.dumps(metadata, indent=4, ensure_ascii=False)
                        outfile.write(json_string)
                    return True, form_id
                else:
                    return False, request.translate("The form already exists in this project")
            else:
                return False, request.translate(
                    "The form ID has especial characters. FormShare only allows letters, numbers and underscores(_)")
        else:
            return False, request.translate(
                "Cannot find XForm ID. Please send this form to support_formshare@qlands.com")
    except PyXFormError as e:
        log.error("Error {} while adding form {} in project {}".format(str(e), input_file_name, project_id))
        return False, str(e)
    except Exception as e:
        log.error("Error {} while adding form {} in project {}".format(str(e), input_file_name, project_id))
        return False, sys.exc_info()[0]


def update_odk_form(request, project_id, for_form_id, user_id, odk_dir, form_data):
    uid = str(uuid.uuid4())
    form_directory = uid[-12:]
    paths = ['tmp', uid]
    os.makedirs(os.path.join(odk_dir, *paths))

    input_file = request.POST['xlsx'].file
    input_file_name = request.POST['xlsx'].filename

    paths = ['tmp', uid, input_file_name]
    file_name = os.path.join(odk_dir, *paths)

    input_file.seek(0)
    with open(file_name, 'wb') as permanent_file:
        shutil.copyfileobj(input_file, permanent_file)

    input_file.close()

    parts = os.path.splitext(input_file_name)

    paths = ['tmp', uid, parts[0]+'.xml']
    xml_file = os.path.join(odk_dir, *paths)

    try:
        xls2xform.xls2xform_convert(file_name, xml_file)

        # Check if the conversion created itemsets.csv
        output_dir = os.path.split(xml_file)[0]
        itemsets_csv = os.path.join(output_dir, "itemsets.csv")
        if not os.path.isfile(itemsets_csv):
            itemsets_csv = ""

        tree = etree.parse(xml_file)
        root = tree.getroot()
        nsmap = root.nsmap[None]
        h_nsmap = root.nsmap['h']
        eid = root.findall(".//{" + nsmap + "}" + parts[0])
        if eid:
            form_id = eid[0].get("id")
            if re.match(r'^[A-Za-z0-9_]+$', form_id):
                if form_id == for_form_id:
                    form_title = root.findall(".//{" + h_nsmap + "}title")
                    if form_exists(request, project_id, form_id):
                        paths = ['forms', form_directory, 'media']
                        if not os.path.exists(os.path.join(odk_dir, *paths)):
                            os.makedirs(os.path.join(odk_dir, *paths))

                            paths = ['forms', form_directory, 'submissions']
                            os.makedirs(os.path.join(odk_dir, *paths))

                            paths = ['forms', form_directory, 'submissions', 'logs']
                            os.makedirs(os.path.join(odk_dir, *paths))

                            paths = ['forms', form_directory, 'submissions', 'maps']
                            os.makedirs(os.path.join(odk_dir, *paths))

                            paths = ['forms', form_directory, 'repository']
                            os.makedirs(os.path.join(odk_dir, *paths))
                            paths = ['forms', form_directory, 'repository', 'temp']
                            os.makedirs(os.path.join(odk_dir, *paths))

                        xls_file_fame = os.path.basename(file_name)
                        xml_file_name = os.path.basename(xml_file)
                        paths = ['forms', form_directory, xls_file_fame]
                        final_xls = os.path.join(odk_dir, *paths)
                        paths = ['forms', form_directory, xml_file_name]
                        final_xml = os.path.join(odk_dir, *paths)
                        shutil.copyfile(file_name, final_xls)
                        shutil.copyfile(xml_file, final_xml)
                        geopoint = get_geopoint_variable_from_xlsx(final_xls)
                        form_data['project_id'] = project_id
                        form_data['form_id'] = form_id
                        form_data['form_name'] = form_title[0].text
                        form_data['form_cdate'] = datetime.datetime.now()
                        form_data['form_directory'] = form_directory
                        form_data['form_accsub'] = 1
                        form_data['form_testing'] = 1
                        form_data['form_xlsfile'] = final_xls
                        form_data['form_xmlfile'] = final_xml
                        form_data['form_pubby'] = user_id
                        form_data['form_geopoint'] = geopoint

                        updated, message = update_form(request, project_id, for_form_id, form_data)
                        if not updated:
                            return updated, message

                        # If we have itemsets.csv add it as media files
                        if itemsets_csv != "":
                            bucket_id = project_id + form_id
                            bucket_id = md5(bucket_id.encode('utf-8')).hexdigest()
                            try:
                                delete_stream(request, bucket_id, 'itemsets.csv')
                            except Exception as e:
                                log.error("Error {} removing filename {} from bucket {}".format(str(e), 'itemsets.csv',
                                                                                                bucket_id))
                            with open(itemsets_csv, "rb", ) as itemset_file:
                                md5sum = md5(itemset_file.read()).hexdigest()
                                added, message = add_file_to_form(request, project_id, form_id, 'itemsets.csv', None, True,
                                                                  md5sum)
                                if added:
                                    itemset_file.seek(0)
                                    store_file(request, bucket_id, 'itemsets.csv', itemset_file)

                        paths = ['forms', form_directory, parts[0]+".json"]
                        json_file = os.path.join(odk_dir, *paths)

                        metadata = {"formID": form_id, "name": form_title[0].text, "majorMinorVersion": "",
                                    "version": datetime.datetime.now().strftime("%Y%m%d"),
                                    "hash": 'md5:' + md5(open(final_xml, 'rb').read()).hexdigest(),
                                    "descriptionText": form_title[0].text}

                        with open(json_file, "w",) as outfile:
                            json_string = json.dumps(metadata, indent=4, ensure_ascii=False)
                            outfile.write(json_string)
                        return True, form_id
                    else:
                        return False, request.translate("The form does not exists in this project")
                else:
                    return False, request.translate('The "form_id" of the current form does not match the "form_id" of '
                                                    'the one you uploaded. You cannot update a form with another form')
            else:
                return False, request.translate(
                    "The form ID has especial characters. FormShare only allows letters, numbers and underscores(_)")
        else:
            return False, request.translate(
                "Cannot find XForm ID. Please send this form to support_formshare@qlands.com")
    except PyXFormError as e:
        log.error("Error {} while adding form {} in project {}".format(str(e), input_file_name, project_id))
        return False, str(e)
    except Exception as e:
        log.error("Error {} while adding form {} in project {}".format(str(e), input_file_name, project_id))
        return False, sys.exc_info()[0]

class FileIterator(object):
    chunk_size = 4096

    def __init__(self, filename):
        self.filename = filename
        self.fileobj = open(self.filename, 'rb')

    def __iter__(self):
        return self

    def next(self):
        chunk = self.fileobj.read(self.chunk_size)
        if not chunk:
            raise StopIteration
        return chunk

    __next__ = next  # py3 compat


# An Object containing the file download iterator
class FileIterable(object):
    def __init__(self, filename):
        self.filename = filename

    def __iter__(self):
        return FileIterator(self.filename)


def generate_form_list(project_array):
    root = etree.Element("xforms", xmlns="http://openrosa.org/xforms/xformsList")
    for project in project_array:
        xform_tag = etree.Element("xform")
        for key, value in project.items():
            atag = etree.Element(key)
            atag.text = value
            xform_tag.append(atag)
        root.append(xform_tag)
    return etree.tostring(root, encoding='utf-8')


def generate_manifest(media_file_array):
    root = etree.Element("manifest", xmlns="http://openrosa.org/xforms/xformsManifest")
    for file in media_file_array:
        xform_tag = etree.Element("mediaFile")
        for key, value in file.items():
            atag = etree.Element(key)
            atag.text = value
            xform_tag.append(atag)
        root.append(xform_tag)
    return etree.tostring(root, encoding='utf-8')


def get_form_list(request, user, project_code, assistant):
    project_id = get_project_id_from_name(request, user, project_code)
    assistant_project = get_project_from_assistant(request, user, project_id, assistant)
    prj_list = []
    odk_dir = get_odk_path(request)
    forms = get_assistant_forms(request, project_id, assistant_project, assistant)
    for form in forms:
        path = os.path.join(odk_dir, *["forms", form["form_directory"], '*.json'])
        files = glob.glob(path)
        if files:
            with io.open(files[0], encoding='utf8') as data_file:
                data = json.load(data_file)
                data["downloadUrl"] = request.route_url('odkxmlform', userid=user, projcode=project_code,
                                                        formid=form["form_id"])
                data["manifestUrl"] = request.route_url('odkmanifest', userid=user, projcode=project_code,
                                                        formid=form["form_id"])
            prj_list.append(data)
    return generate_form_list(prj_list)


def get_manifest(request, user, project, project_id, form):
    form_files = get_form_files(request, project_id, form)
    for file in form_files:
        if file['file_url'] is not None:
            cache_external_file(request, project_id, form, file['file_name'])

    form_files = get_form_files(request, project_id, form)
    if form_files:
        file_array = []
        for file in form_files:
            file_name = file['file_name']
            file_array.append({'filename': file_name, 'hash': 'md5:' + file['file_md5'],
                               'downloadUrl': request.route_url('odkmediafile', userid=user, projcode=project,
                                                                formid=form, fileid=file_name)})
        return generate_manifest(file_array)
    else:
        return generate_manifest([])


def get_xml_form(request, project, form):
    xml_file = get_form_xml_file(request, project, form)
    if xml_file is not None:
        content_type, content_enc = mimetypes.guess_type(xml_file)
        file_name = os.path.basename(xml_file)
        response = FileResponse(
            xml_file,
            request=request,
            content_type=content_type
        )
        response.content_disposition = 'attachment; filename="' + file_name + '"'
        return response
    else:
        raise HTTPNotFound()


def get_media_file(request, project_id, form_id, file_id):
    return retrieve_form_file(request, project_id, form_id, file_id)


def get_submission_file(request, project, form, submission):
    odk_dir = get_odk_path(request)
    form_directory = get_form_directory(request, project, form)
    if form_directory is not None:
        path = os.path.join(odk_dir, *["forms", form_directory, 'submissions', submission + ".json"])
        if os.path.isfile(path):
            content_type, content_enc = mimetypes.guess_type(path)
            file_name = os.path.basename(path)
            response = FileResponse(
                path,
                request=request,
                content_type=content_type
            )
            response.content_disposition = 'attachment; filename="' + file_name + '"'
            return response
        else:
            raise HTTPNotFound()
    else:
        raise HTTPNotFound()


class ChangeDir:
    def __init__(self, new_path):
        self.newPath = os.path.expanduser(new_path)

    # Change directory with the new path
    def __enter__(self):
        self.savedPath = os.getcwd()
        os.chdir(self.newPath)

    # Return back to previous directory
    def __exit__(self, etype, value, traceback):
        os.chdir(self.savedPath)


def build_odata_service(request, schema):
    instance_id = str(uuid.uuid4())
    instance_id = instance_id[-12:]
    temp_dir = os.path.join(get_odk_path(request), *["tmp", instance_id])
    os.makedirs(temp_dir)
    yml_file = os.path.join(get_odk_path(request), *["tmp", instance_id, schema + ".yml"])
    file = open(yml_file, "w")
    file.write("connector:\n")
    file.write("  mysql:\n")
    file.write("    user: " + request.registry.settings['mysql.user'] + "\n")
    file.write("    password: " + request.registry.settings['mysql.password'] + "\n")
    file.write("    host: " + request.registry.settings['mysql.host'] + "\n")
    file.write("    port: " + request.registry.settings['mysql.port'] + "\n")
    file.write("    buffered: True\n")
    file.write("generator:\n")
    file.write("  group_id: org.cabi." + schema + "\n")
    file.write("  artifact_id: " + schema + "\n")
    file.write("  project_description: CABI OData " + schema + "\n")
    file.write("  project_name: " + schema + "\n")
    file.write("  root_package_name: org.cabi." + schema + "\n")
    file.close()
    args = ["odata_generator", "-c", yml_file, schema, temp_dir]

    try:
        p = Popen(args, stdout=PIPE, stderr=PIPE)
        stdout, stderr = p.communicate()
        if p.returncode == 0:
            with ChangeDir(temp_dir):
                args = ["mvn", "clean", "package"]
                p = Popen(args, stdout=PIPE, stderr=PIPE)
                stdout, stderr = p.communicate()
                if p.returncode != 0:
                    log.error("Error creating package with maven")
                    log.error("In directory: " + os.getcwd())
                    log.error("Error:" + stdout.decode() + " - " + stderr.decode() + " - " + " ".join(args))
                else:
                    source_war_file = os.path.join(get_odk_path(request),
                                                   *["tmp", instance_id, "target", schema + ".war"])
                    target_war_file = os.path.join(request.registry.settings['tomcat.wardirectory'], schema + ".war")
                    try:
                        shutil.copyfile(source_war_file, target_war_file)
                    except Exception as e:
                        log.error("Unable to copy WAR file to final directory:" + str(e))
        else:
            log.error("ODATA Generator Error: " + stdout.decode() + " - " + stderr.decode() + " - " + " ".join(args))
    
        # Do not stop the process if OData does not build
        return 0, ""
    except Exception as e:
        log.error("Error generating ODK: " + str(e))
        return 0, ""


def build_database(cnf_file, create_file, insert_file, audit_file, schema):
    error = False
    error_message = ""

    args = ["mysql", "--defaults-file=" + cnf_file, '--execute=DROP SCHEMA IF EXISTS ' + schema]
    try:
        check_call(args)
    except CalledProcessError as e:
        error_message = "Error dropping schema \n"
        error_message = error_message + "Error: \n"
        if e.stderr is not None:
            error_message = error_message + e.stderr.encode() + "\n"
        log.error(error_message)
        error = True

    if not error:
        args = ["mysql", "--defaults-file=" + cnf_file,
                '--execute=CREATE SCHEMA ' + schema + ' DEFAULT CHARACTER SET utf8 DEFAULT COLLATE utf8_general_ci']
        try:
            check_call(args)
        except CalledProcessError as e:
            error_message = "Error dropping schema \n"
            error_message = error_message + "Error: \n"
            if e.stderr is not None:
                error_message = error_message + e.stderr.encode() + "\n"
            log.error(error_message)
            error = True

    if not error:
        args = ["mysql", "--defaults-file=" + cnf_file, schema]

        with open(create_file) as input_file:
            proc = Popen(
                args, stdin=input_file, stderr=PIPE, stdout=PIPE)
            output, error = proc.communicate()
            if proc.returncode != 0:
                error_message = "Error creating database \n"
                error_message = error_message + "File: " + create_file + "\n"
                error_message = error_message + "Error: \n"
                error_message = error_message + error.encode() + "\n"
                error_message = error_message + "Output: \n"
                error_message = error_message + output.encode() + "\n"
                log.error(error_message)
                error = True

    if not error:
        with open(insert_file) as input_file:
            proc = Popen(args, stdin=input_file, stderr=PIPE, stdout=PIPE)
            output, error = proc.communicate()
            if proc.returncode != 0:
                error_message = "Error loading lookup tables \n"
                error_message = error_message + "File: " + insert_file + "\n"
                error_message = error_message + "Error: \n"
                error_message = error_message + error.encode() + "\n"
                error_message = error_message + "Output: \n"
                error_message = error_message + output.encode() + "\n"
                log.error(error_message)
                error = True

    if not error:
        with open(audit_file) as input_file:
            proc = Popen(args, stdin=input_file, stderr=PIPE, stdout=PIPE)
            output, error = proc.communicate()
            if proc.returncode != 0:
                error_message = "Error loading lookup tables \n"
                error_message = error_message + "File: " + audit_file + "\n"
                error_message = error_message + "Error: \n"
                error_message = error_message + error.encode() + "\n"
                error_message = error_message + "Output: \n"
                error_message = error_message + output.encode() + "\n"
                log.error(error_message)
                error = True

    return error, error_message


def create_repository(request, project, form, odk_dir, xform_directory, primary_key, separation_file=None,
                      default_language=None, other_languages=None, yes_no_strings=None):
    print("*************************** create_repository ***********************")
    odk_to_mysql = os.path.join(request.registry.settings['odktools.path'], *["ODKToMySQL", "odktomysql"])

    xlsx_file = get_form_xls_file(request, project, form)
    if xlsx_file is not None:
        args = [odk_to_mysql, "-x " + xlsx_file, "-t maintable", "-v " + primary_key,
                "-c " + os.path.join(odk_dir, *["forms", xform_directory, "repository", "create.sql"]),
                "-C " + os.path.join(odk_dir, *["forms", xform_directory, "repository", "create.xml"]),
                "-i " + os.path.join(odk_dir, *["forms", xform_directory, "repository", "insert.sql"]),
                "-D " + os.path.join(odk_dir, *["forms", xform_directory, "repository", "drop.sql"]),
                "-I " + os.path.join(odk_dir, *["forms", xform_directory, "repository", "insert.xml"]),
                "-m " + os.path.join(odk_dir, *["forms", xform_directory, "repository", "metadata.sql"]),
                "-f " + os.path.join(odk_dir, *["forms", xform_directory, "repository", "manifest.xml"])]

        if separation_file is not None:
            args.append("-s " + os.path.join(odk_dir, *["forms", xform_directory, "repository", separation_file]))
        args.append("-S " + os.path.join(odk_dir, *["forms", xform_directory, "repository", "separation.xml"]))
        if other_languages is not None:
            args.append("-l '" + other_languages + "'")
        if default_language is not None:
            args.append("-d '" + default_language + "'")
        args.append("-T " + os.path.join(odk_dir, *["forms", xform_directory, "repository", "iso639.sql"]))
        if yes_no_strings == '':
            yes_no_strings = None
        if yes_no_strings is not None:
            args.append("-y '" + yes_no_strings + "'")
        args.append("-e " + os.path.join(odk_dir, *["forms", xform_directory, "repository", "temp"]))
        args.append("-o m")

        # Append all media files

        media_files = get_media_files(request, project, form)
        tmp_uid = str(uuid.uuid4())
        target_dir = os.path.join(odk_dir, *["tmp", tmp_uid])
        os.makedirs(target_dir)
        for media_file in media_files:
            store_file_in_directory(request, project, form, media_file.file_name, target_dir)
        path = os.path.join(odk_dir, *["tmp", tmp_uid, '*.*'])
        files = glob.glob(path)
        if files:
            for aFile in files:
                args.append(aFile)

        print("*****************************100")
        print(" ".join(args))
        print("*****************************100")

        p = Popen(args, stdout=PIPE, stderr=PIPE)
        stdout, stderr = p.communicate()
        if p.returncode == 0:
            schema = str(uuid.uuid4())
            schema = "D" + schema[-12:]
            create_file = os.path.join(odk_dir, *["forms", xform_directory, "repository", "create.sql"])
            insert_file = os.path.join(odk_dir, *["forms", xform_directory, "repository", "insert.sql"])
            audit_file = os.path.join(odk_dir, *["forms", xform_directory, "repository", "audit.sql"])

            audit = open(audit_file, "w")
            audit.write("CREATE TABLE IF NOT EXISTS audit_log (\n")
            audit.write("audit_id VARCHAR(64) NOT NULL,\n")
            audit.write("audit_date DATETIME NULL ,\n")
            audit.write("audit_action VARCHAR(6) NULL ,\n")
            audit.write("audit_user VARCHAR(120) NULL ,\n")
            audit.write("audit_table VARCHAR(120) NULL ,\n")
            audit.write("audit_column VARCHAR(120) NULL ,\n")
            audit.write("audit_uuid VARCHAR(64) NULL ,\n")
            audit.write("audit_oldvalue TEXT NULL ,\n")
            audit.write("audit_newvalue TEXT NULL ,\n")
            audit.write("audit_insdeldata TEXT NULL ,\n")
            audit.write("PRIMARY KEY (audit_id) )\n")
            audit.write(" ENGINE = InnoDB CHARSET=utf8;\n")
            audit.close()

            error, db_message = build_database(request.registry.settings['mysql.cnf'], create_file, insert_file,
                                               audit_file, schema)
            if not error:
                odata_error_code, odata_error_message = build_odata_service(request, schema)
                if odata_error_code == 0:
                    form_data = {'form_schema': schema, 'form_pkey': primary_key}
                    update_form(request, project, form, form_data)

                    # Remove any test submissions if any. In try because nothing happens if they
                    # don't get removed... just junk files
                    submissions_path = os.path.join(odk_dir, *['forms', xform_directory, "submissions", '*.*'])
                    files = glob.glob(submissions_path)
                    if files:
                        for file in files:
                            try:
                                os.remove(file)
                            except Exception as e:
                                log.error(str(e))
                    submissions_path = os.path.join(odk_dir, *['forms', xform_directory, "submissions", '*/'])
                    files = glob.glob(submissions_path)
                    if files:
                        for file in files:
                            if file[-5:] != "logs/" and file[-5:] != "maps/":
                                try:
                                    shutil.rmtree(file)
                                except Exception as e:
                                    log.error(str(e))

                    return 0, ""
                else:
                    return 1, odata_error_message
            else:
                return 1, db_message
        else:
            if p.returncode == 1 or p.returncode < 0:
                return p.returncode, stdout.decode() + " - " + stderr.decode() + " - " + " ".join(args)
            else:
                return p.returncode, stdout.decode()
    else:
        return -1, "Form data cannot be found"


def move_media_files(odk_dir, xform_directory, src_submission, trg_submission):
    source_path = os.path.join(odk_dir, *['forms', xform_directory, "submissions", src_submission, '*.*'])
    target_path = os.path.join(odk_dir, *['forms', xform_directory, "submissions", trg_submission])
    files = glob.glob(source_path)
    for file in files:
        try:
            shutil.move(file, target_path)
        except Exception as e:
            log.error(
                "moveMediaFiles. Error moving from " + src_submission + " to " + trg_submission + " . Message: " + str(
                    e))


def convert_xml_to_json(odk_dir, xml_file, xform_directory, schema, xml_form_file, user, project, form, assistant,
                        request):
    xml_to_json = os.path.join(request.registry.settings['odktools.path'], *["XMLtoJSON", "xmltojson"])
    temp_json_file = xml_file.replace(".xml", ".tmp.json")
    json_file = xml_file.replace(".xml", ".json")
    submission_id = os.path.basename(xml_file)
    submission_id = submission_id.replace(".xml", "")

    # First we convert the XML to a temporal JSON
    args = [xml_to_json, "-i " + xml_file, "-o " + temp_json_file, "-x " + xml_form_file]
    p = Popen(args, stdout=PIPE, stderr=PIPE)
    stdout, stderr = p.communicate()
    if p.returncode == 0:
        if schema is not None:
            if schema != "":
                # Add the controlling fields to the JSON file
                project_code = get_project_code_from_id(request, user, project)
                with open(temp_json_file, 'r') as f:
                    submission_data = json.load(f)
                    submission_data["_submitted_by"] = assistant
                    submission_data["_submitted_date"] = datetime.datetime.now().isoformat()
                    submission_data["_user_id"] = user
                    submission_data["_submission_id"] = submission_id
                    submission_data["_project_code"] = project_code
                    geopoint_variable = get_form_geopoint(request, project, form)
                    if geopoint_variable is not None:
                        if geopoint_variable in submission_data.keys():
                            submission_data["_geopoint"] = submission_data[geopoint_variable]

                with open(temp_json_file, "w", ) as outfile:
                    json_string = json.dumps(submission_data, indent=4, ensure_ascii=False)
                    outfile.write(json_string)

                # Second we pass the temporal JSON to jQ to order its elements
                # this will help later on if we want to compare between JSONs
                args = ["jq", "-S", ".", temp_json_file]

                final = open(json_file, "w")
                p = Popen(args, stdout=final, stderr=PIPE)
                p.communicate()
                final.close()
                if p.returncode == 0:
                    try:
                        os.remove(temp_json_file)
                    except Exception as e:
                        log.error(
                            "XMLToJSON error. Temporary file " + temp_json_file + " might not exist! Error: " + str(e))
                        return 1, ""

                    # Get the MD5Sum of the JSON file and looks for it in the submissions
                    # This will get an exact match that will not move into the database
                    md5sum = md5(open(json_file, 'rb').read()).hexdigest()
                    sameas = get_submission_data(request, project, form, md5sum)
                    if sameas is None:
                        # Third we try to move the JSON data into the database
                        mysql_user = request.registry.settings['mysql.user']
                        mysql_password = request.registry.settings['mysql.password']
                        mysql_host = request.registry.settings['mysql.host']
                        mysql_port = request.registry.settings['mysql.port']
                        log_file = os.path.join(odk_dir, *['forms', xform_directory, 'submissions', 'logs',
                                                           os.path.basename(xml_file)])
                        imported_file = os.path.join(odk_dir,
                                                     *['forms', xform_directory, 'submissions', 'logs', 'imported.log'])
                        maps_directory = os.path.join(odk_dir, *['forms', xform_directory, 'submissions', 'maps'])
                        manifest_file = os.path.join(odk_dir, *['forms', xform_directory, 'repository', 'manifest.xml'])
                        args = []
                        json_to_mysql = os.path.join(request.registry.settings['odktools.path'],
                                                     *["JSONToMySQL", "jsontomysql"])
                        args.append(json_to_mysql)
                        args.append("-H " + mysql_host)
                        args.append("-P " + mysql_port)
                        args.append("-u " + mysql_user)
                        args.append("-p " + mysql_password)
                        args.append("-s " + schema)
                        args.append("-o " + log_file)
                        args.append("-j " + json_file)
                        args.append("-i " + imported_file)
                        args.append("-M " + maps_directory)
                        args.append("-m " + manifest_file)
                        args.append("-O m")
                        args.append("-w")
                        p = Popen(args, stdout=PIPE, stderr=PIPE)
                        stdout, stderr = p.communicate()
                        # An error 2 is an SQL error that goes to the logs
                        if p.returncode == 0 or p.returncode == 2:
                            project_of_assistant = get_project_from_assistant(request, user, project, assistant)
                            added, message = add_submission(request, project, form, project_of_assistant, assistant,
                                                            submission_id, md5sum, p.returncode)

                            if not added:
                                log.error(message)
                                return 1, message

                            if p.returncode == 2:
                                added, message = add_json_log(request, project, form, submission_id, json_file,
                                                              log_file, 1, project_of_assistant, assistant)
                                if not added:
                                    log.error(message)
                                    return 1, message
                            else:
                                # Add the JSON to the Elastic Search index but only submissions without error
                                create_dataset_index(request, user, project_code, form)
                                add_dataset(request, user, project_code, form, submission_id, submission_data)

                            return 0, ""
                        else:
                            log.error(
                                "JSONToMySQL error. Inserting " + json_file + ". Error: " + stdout.decode() + "-" +
                                stderr.decode() + ". Command line: " + " ".join(args))
                            return 2, ""
                    else:
                        # If the MD5Sum is the same then add it to the submission table
                        # indicating the "sameas" field. Any media files of the
                        # duplicated submission moves to the "sameas" submission.
                        # This will fix the issue when the media files are so big
                        # that multiple posts are done by ODK Collect for the same
                        # submission
                        project_of_assistant = get_project_from_assistant(request, user, project, assistant)
                        added, message = add_submission_same_as(request, project, form, project_of_assistant, assistant,
                                                                submission_id, md5sum, 0, sameas.submission_id)
                        if not added:
                            log.error(message)
                            return 1, message

                        move_media_files(odk_dir, xform_directory, submission_id, sameas.submission_id)
                        return 0, ""
                else:
                    log.error(
                        "jQ error. Converting " + xml_file + "  to " + json_file + ". Error: " + "-" +
                        stderr.decode() + ". Command line: " + " ".join(args))
                    return 1, ""
        else:
            # We compare the MD5Sum of the testing submissions so the
            # media files are stored in the proper way if ODK Collect
            # send the media files in separate submissions when such
            # file are so big that separation is required
            try:
                shutil.move(temp_json_file, json_file)
                project_code = get_project_code_from_id(request, user, project)
                with open(json_file, 'r') as f:
                    submission_data = json.load(f)
                    submission_data["_submitted_by"] = assistant
                    submission_data["_submitted_date"] = datetime.datetime.now().isoformat()
                    submission_data["_user_id"] = user
                    submission_data["_submission_id"] = submission_id
                    submission_data["_project_code"] = project_code
                    geopoint_variable = get_form_geopoint(request, project, form)
                    if geopoint_variable is not None:
                        try:
                            submission_data["_geopoint"] = submission_data[geopoint_variable]
                        except KeyError:
                            pass

                # Adds the dataset to the Elastic Search index
                create_dataset_index(request, user, project_code, form)
                add_dataset(request, user, project_code, form, submission_id, submission_data)

                with open(json_file, "w", ) as outfile:
                    json_string = json.dumps(submission_data, indent=4, ensure_ascii=False)
                    outfile.write(json_string)

                submissions_path = os.path.join(odk_dir, *['forms', xform_directory, "submissions", '*.json'])
                files = glob.glob(submissions_path)
                md5sum = md5(open(json_file, 'rb').read()).hexdigest()
                for aFile in files:
                    if aFile != json_file:
                        othmd5sum = md5(open(aFile, 'rb').read()).hexdigest()
                        if md5sum == othmd5sum:
                            target_submission_id = os.path.basename(aFile)
                            target_submission_id = target_submission_id.replace(".json", "")
                            move_media_files(odk_dir, xform_directory, target_submission_id, submission_id)
                            os.remove(aFile)
                return 0, ""
            except Exception as e:
                log.error("XMLToJSON error. Temporary file " + temp_json_file + " might not exist! Error: " + str(e))
                return 1, ""
    else:
        log.error("XMLToJSON error. Converting " + xml_file + "  to " + json_file + ". Error: " + stdout.decode() +
                  "-" + stderr.decode() + ". Command line: " + " ".join(args))
        return 1, ""


def store_submission(request, user, project, assistant):
    odk_dir = get_odk_path(request)
    unique_id = uuid4()
    path = os.path.join(odk_dir, *["submissions", str(unique_id)])
    os.makedirs(path)
    xml_file = ""
    for key in request.POST.keys():
        try:
            filename = request.POST[key].filename
            if filename.upper().find('.XML') >= 0:
                filename = str(unique_id) + ".xml"
            input_file = request.POST[key].file
            file_path = os.path.join(path, filename)
            if file_path.upper().find('.XML') >= 0:
                xml_file = file_path
            temp_file_path = file_path + '~'
            input_file.seek(0)
            with open(temp_file_path, 'wb') as output_file:
                shutil.copyfileobj(input_file, output_file)
            # Now that we know the file has been fully saved to disk move it into place.
            os.rename(temp_file_path, file_path)
        except Exception as e:
            log.error("Submission " + str(unique_id) + " has POST error key: " + key + " Error: " + str(e) +
                      ". URL: " + request.url)
    if xml_file != "":
        tree = etree.parse(xml_file)
        root = tree.getroot()
        xform_id = root.get("id")
        if xform_id is not None:
            form_data = get_form_data(request, project, xform_id)
            if form_data is not None:
                if assistant_has_form(request, user, project, xform_id, assistant):
                    media_path = os.path.join(odk_dir,
                                              *['forms', form_data['form_directory'], "submissions", str(unique_id),
                                                'diffs'])
                    os.makedirs(media_path)
                    media_path = os.path.join(odk_dir,
                                              *['forms', form_data['form_directory'], "submissions", str(unique_id)])
                    target_path = os.path.join(odk_dir, *['forms', form_data['form_directory'], "submissions"])
                    path = os.path.join(path, *['*.*'])
                    files = glob.glob(path)
                    xml_file = ""
                    for file in files:
                        base_file = os.path.basename(file)
                        if base_file.upper().find('.XML') >= 0:
                            target_file = os.path.join(target_path, base_file)
                            xml_file = target_file
                        else:
                            target_file = os.path.join(media_path, base_file)
                        shutil.move(file, target_file)
                    if xml_file != "":
                        res_code, message = convert_xml_to_json(odk_dir, xml_file, form_data['form_directory'],
                                                                form_data['form_schema'], form_data['form_xmlfile'],
                                                                user, project, xform_id, assistant, request)
                        if res_code == 0:
                            return True, 201
                        else:
                            if res_code == 1:
                                return False, 500
                            else:
                                return True, 201
                    else:
                        return False, 404
                else:
                    log.error('Enumerator %s cannot submit data to %s', assistant, xform_id)
                    return False, 404
            else:
                log.error('Submission for ID %s does not exist in the database', xform_id)
                return False, 404
        else:
            log.error('Submission does not have and ID')
            return False, 404
    else:
        log.error('Submission does not have an XML file')
        return False, 500


def get_html_from_diff(request, project, form, submission, revision):
    odk_dir = get_odk_path(request)
    form_directory = get_form_directory(request, project, form)

    diff_file = os.path.join(odk_dir, *["forms", form_directory, "submissions", submission, "diffs", revision,
                                        revision + '.diff'])
    html_file = os.path.join(odk_dir, *["forms", form_directory, "submissions", submission, "diffs", revision,
                                        revision + '.html'])

    if not os.path.exists(html_file):
        args = ["diff2html", "-s", "side", "-o", "stdout", "-i", "file", diff_file]

        final = open(html_file, "w")
        p = Popen(args, stdout=final, stderr=PIPE)
        nada, stderr = p.communicate()
        final.close()
        if p.returncode == 0:
            soup = BeautifulSoup(open(html_file), "html.parser")
            for diff in soup.find_all('div', {'class': 'd2h-files-diff'}):
                return 0, diff
            return 0, ""
        else:
            message = "Generating HTML " + html_file + ". Error: " + "-" + stderr.decode() + ". Command line: " \
                      + " ".join(args)
            log.error(message)
            return 1, message
    else:
        soup = BeautifulSoup(open(html_file), "html.parser")
        for diff in soup.find_all('div', {'class': 'd2h-files-diff'}):
            return 0, diff


def generate_diff(request, project, form, json_file_a, json_file_b):
    odk_dir = get_odk_path(request)
    form_directory = get_form_directory(request, project, form)
    file_a = os.path.join(odk_dir, *['forms', form_directory, 'submissions', json_file_a + '.json'])
    file_b = os.path.join(odk_dir, *['forms', form_directory, 'submissions', json_file_b + '.json'])

    diff_id = str(uuid.uuid4())
    diff_id = "DIFF" + diff_id[-12:]
    temp_dir = os.path.join(odk_dir, *["tmp", diff_id])
    os.makedirs(temp_dir)
    diff_file = os.path.join(odk_dir, *["tmp", diff_id, diff_id + ".diff"])
    html_file = os.path.join(odk_dir, *["tmp", diff_id, diff_id + ".html"])

    args = ["diff", "-u", file_a, file_b]

    final = open(diff_file, "w")
    p = Popen(args, stdout=final, stderr=PIPE)
    nada, stderr = p.communicate()
    final.close()
    if p.returncode == 1:
        args = ["diff2html", "-s", "side", "-o", "stdout", "-i", "file", diff_file]

        final = open(html_file, "w")
        p = Popen(args, stdout=final, stderr=PIPE)
        nada, stderr = p.communicate()
        final.close()
        if p.returncode == 0:
            soup = BeautifulSoup(open(html_file), "html.parser")
            for diff in soup.find_all('div', {'class': 'd2h-files-diff'}):
                return 0, diff
            return 0, ""
        else:
            message = "Generating HTML " + html_file + ". Error: " + "-" + stderr.decode() + ". Command line: " + \
                      " ".join(args)
            log.error(message)
            return 1, message
    else:
        if p.returncode != 0:
            message = "DIFF Error. Comparing " + file_a + "  to " + file_b + ". Error: " + "-" + stderr.decode() + \
                      ". Command line: " + " ".join(args)
            log.error(message)
            return 1, message
        else:
            return 1, "The files are the same"


def store_new_version(request, user, project, form, submission, assistant, sequence, new_file, notes):
    odk_dir = get_odk_path(request)
    form_directory = get_form_directory(request, project, form)
    diff_path = os.path.join(odk_dir, *['forms', form_directory, 'submissions', submission, 'diffs', sequence])
    try:
        os.makedirs(diff_path)
        current_file = os.path.join(odk_dir, *['forms', form_directory, 'submissions', submission + ".json"])
        backup_file = os.path.join(odk_dir, *['forms', form_directory, 'submissions', submission, 'diffs', sequence,
                                              submission + ".json"])
        diff_file = os.path.join(odk_dir, *['forms', form_directory, 'submissions', submission, 'diffs', sequence,
                                            sequence + ".diff"])
        try:
            shutil.copyfile(current_file, backup_file)
            temp_file_path = current_file + '~'
            try:
                with open(temp_file_path, 'wb') as output_file:
                    shutil.copyfileobj(new_file, output_file)
                os.remove(current_file)
                os.rename(temp_file_path, current_file)

                args = ["diff", "-u", backup_file, current_file]

                final = open(diff_file, "w")
                p = Popen(args, stdout=final, stderr=PIPE)
                nada, stderr = p.communicate()
                final.close()
                if p.returncode == 1:
                    update_json_status(request, project, form, submission, 3)
                    project_of_assistant = get_project_from_assistant(request, user, project, assistant)
                    added, message = add_json_history(request, project, form, submission, sequence, 3,
                                                      project_of_assistant, assistant, notes)
                    if not added:
                        return 1, message

                    return 0, ""
                else:
                    shutil.copyfile(backup_file, current_file)
                    if p.returncode != 0:
                        message = "DIFF Error. Comparing " + current_file + "  to " + backup_file + ". Error: " + \
                                  "-" + stderr.decode() + ". Command line: " + " ".join(args)
                        log.error(message)
                        return 1, message
                    else:
                        return 1, "The files are the same"
            except Exception as e:
                log.error(str(e))
                if not os.path.exists(current_file):
                    shutil.copyfile(backup_file, current_file)
                return 1, "Cannot save new submission"
        except Exception as e:
            log.error(str(e))
            return 1, "Cannot create backup of current submission"
    except Exception as e:
        log.error(str(e))
        return 1, "Cannot created path for submission"


def restore_from_revision(request, project, form, submission, sequence):
    dok_dir = get_odk_path(request)
    form_directory = get_form_directory(request, project, form)
    current_file = os.path.join(dok_dir, *['forms', form_directory, 'submissions', submission + ".json"])
    backup_file = os.path.join(dok_dir, *['forms', form_directory, 'submissions', submission, 'diffs', sequence,
                                          submission + ".json"])
    try:
        shutil.copyfile(backup_file, current_file)
        return 0, ""
    except Exception as e:
        log.error(str(e))
        return 1, "Cannot restore from revision " + sequence


def push_revision(request, user, project, form, submission):
    odk_dir = get_odk_path(request)
    form_directory = get_form_directory(request, project, form)
    schema = get_form_schema(request, project, form)
    current_file = os.path.join(odk_dir, *['forms', form_directory, 'submissions', submission + ".json"])

    # Third we try to move the JSON data into the database
    mysql_user = request.registry.settings['mysql.user']
    mysql_password = request.registry.settings['mysql.password']
    mysql_host = request.registry.settings['mysql.host']
    mysql_port = request.registry.settings['mysql.port']
    log_file = os.path.join(odk_dir, *['forms', form_directory, 'submissions', 'logs', submission + ".xml"])
    imported_file = os.path.join(odk_dir, *['forms', form_directory, 'submissions', 'logs', 'imported.log'])
    maps_directory = os.path.join(odk_dir, *['forms', form_directory, 'submissions', 'maps'])
    manifest_file = os.path.join(odk_dir, *['forms', form_directory, 'repository', 'manifest.xml'])
    args = []
    json_to_mysql = os.path.join(request.registry.settings['odktools.path'], *["JSONToMySQL", "jsontomysql"])
    args.append(json_to_mysql)
    args.append("-H " + mysql_host)
    args.append("-P " + mysql_port)
    args.append("-u " + mysql_user)
    args.append("-p " + mysql_password)
    args.append("-s " + schema)
    args.append("-o " + log_file)
    args.append("-j " + current_file)
    args.append("-i " + imported_file)
    args.append("-M " + maps_directory)
    args.append("-m " + manifest_file)
    args.append("-O m")
    args.append("-w")
    p = Popen(args, stdout=PIPE, stderr=PIPE)
    stdout, stderr = p.communicate()
    if p.returncode == 0:
        with open(current_file, 'r') as f:
            submission_data = json.load(f)

            # Add the JSON to the Elastic Search index
            project_code = get_project_code_from_id(request, user, project)
            create_dataset_index(request, user, project_code, form)
            add_dataset(request, user, project_code, form, submission, submission_data)
        return 0, ""
    else:
        log.error("JSONToMySQL error. Pushing " + current_file + ". Error: " + stdout.decode() + "-" +
                  stderr.decode() + ". Command line: " + " ".join(args))
        return 2, ""


def get_tables_from_form(request, project, form):
    odk_dir = get_odk_path(request)
    form_directory = get_form_directory(request, project, form)
    create_file = os.path.join(odk_dir, *['forms', form_directory, 'repository', 'create.xml'])
    tree = etree.parse(create_file)
    root = tree.getroot()
    element_lkp_tables = root.find(".//lkptables")
    element_tables = root.find(".//tables")
    # Append all tables
    tables = element_tables.findall(".//table")
    result = []
    if tables:
        for table in tables:
            result.append({'name': table.get('name'), 'desc': table.get('desc')})
    # Append all lookup tables
    tables = element_lkp_tables.findall(".//table")
    if tables:
        for table in tables:
            result.append({'name': table.get('name'), 'desc': table.get('desc')})
    return result


def get_fields_from_table(request, project, form, table_name, current_fields):
    odk_dir = get_odk_path(request)
    form_directory = get_form_directory(request, project, form)
    create_file = os.path.join(odk_dir, *['forms', form_directory, 'repository', 'create.xml'])
    tree = etree.parse(create_file)
    root = tree.getroot()
    table = root.find(".//table[@name='" + table_name + "']")
    result = []
    checked = 0
    if table is not None:
        for field in table.getchildren():
            if field.tag == "field":
                if field.get('name') != 'rowuuid':
                    found = False
                    for cfield in current_fields:
                        if field.get('name') == cfield:
                            found = True
                            checked = checked + 1
                    desc = field.get('desc')
                    if desc == "":
                        desc = field.get('name') + " - Without description"
                    result.append({'name': field.get('name'), 'desc': desc,
                                   'type': field.get('type'), 'size': field.get('size'),
                                   'decsize': field.get('decsize'), 'checked': found})
            else:
                break
    return result, checked


def get_table_desc(request, project, form, table_name):
    odk_dir = get_odk_path(request)
    form_directory = get_form_directory(request, project, form)
    create_file = os.path.join(odk_dir, *['forms', form_directory, 'repository', 'create.xml'])
    tree = etree.parse(create_file)
    root = tree.getroot()
    table = root.find(".//table[@name='" + table_name + "']")
    if table is not None:
        return table.get("desc")
    return ""


def is_field_key(request, project, form, table_name, field_name):
    if table_name == "maintable" and field_name == "surveyid":
        return True
    if table_name == "maintable" and field_name == "originid":
        return True
    odk_dir = get_odk_path(request)
    form_directory = get_form_directory(request, project, form)
    create_file = os.path.join(odk_dir, *['forms', form_directory, 'repository', 'create.xml'])
    tree = etree.parse(create_file)
    root = tree.getroot()
    table = root.find(".//table[@name='" + table_name + "']")
    if table is not None:
        for field in table.getchildren():
            if field.tag == "field":
                if field.get("name") == field_name:
                    if field.get("key") == "true":
                        return True
                    if field.get("odktype") == "geopoint":
                        return True
            else:
                break
    return False


def get_request_data(request, project, form, table_name, draw, fields, start, length, order_index, order_direction,
                     search_value):
    schema = get_form_schema(request, project, form)
    sql_fields = ','.join(fields)
    not_null_fields_array = []
    for a_field in fields:
        not_null_fields_array.append("IFNULL(" + a_field + ",'')")
    not_null_fields = ','.join(not_null_fields_array)
    table_order = fields[order_index]

    if search_value == "":
        sql = "SELECT " + sql_fields + " FROM " + schema + "." + table_name
        where_clause = ''
    else:
        sql = "SELECT " + sql_fields + " FROM " + schema + "." + table_name
        sql = sql + " WHERE LOWER(CONCAT(" + not_null_fields + ")) like '%" + search_value.lower() + "%'"
        where_clause = " WHERE LOWER(CONCAT(" + not_null_fields + ")) like '%" + search_value.lower() + "%'"

    sql = sql + " ORDER BY " + table_order + " " + order_direction
    sql = sql + " LIMIT " + str(start) + "," + str(length)
    count_sql = "SELECT count(*) as total FROM " + schema + "." + table_name + where_clause
    mark_changed(request.dbsession)
    records = request.dbsession.execute(sql).fetchall()
    data = []
    if records is not None:
        for record in records:
            a_record = {'DT_RowId': record.rowuuid}
            for field in fields:
                try:
                    if isinstance(record[field], datetime.datetime) or isinstance(record[field],
                                                                                  datetime.date) or isinstance(
                            record[field], datetime.time):
                        a_record[field] = record[field].isoformat().replace("T", " ")
                    else:
                        if isinstance(record[field], float):
                            a_record[field] = str(record[field])
                        else:
                            if isinstance(record[field], Decimal):
                                a_record[field] = str(record[field])
                            else:
                                if isinstance(record[field], datetime.timedelta):
                                    a_record[field] = str(record[field])
                                else:
                                    if field != "rowuuid":
                                        a_record[field] = record[field]
                                    else:
                                        a_record[field] = record[field][-12:]
                except Exception as e:
                    a_record[field] = "AJAX Data error. Report this error to support_for_cabi@qlands.com"
                    log.error("AJAX Error in field " + field + ". Error: " + str(e))
            data.append(a_record)

    records = request.dbsession.execute(count_sql).fetchone()
    total = records.total

    result = {'draw': draw, 'recordsTotal': total, 'recordsFiltered': total, 'data': data}
    return result


def update_data(request, project, form, table_name, row_uuid, field, value):
    schema = get_form_schema(request, project, form)
    sql = "UPDATE " + schema + "." + table_name + " SET " + field + " = '" + value + "'"
    sql = sql + " WHERE rowuuid = '" + row_uuid + "'"
    try:
        request.dbsession.execute(sql)
        mark_changed(request.dbsession)
        res = {"data": {field: value}}
        return res
    except exc.IntegrityError as e:
        p1 = re.compile(r'`(\w+)`')
        m1 = p1.findall(str(e))
        if m1:
            if len(m1) == 6:
                lookup = get_table_desc(request, project, form, m1[4])
                return {"fieldErrors": [{'name': field,
                                         'status': 'Cannot update value. Check the valid values in lookup table "' +
                                                   lookup + '"'}]}
        return {
            "fieldErrors": [{'name': field, 'status': 'Cannot update value. Check the valid values in lookup table'}]}
    except Exception as ex:
        log.error(str(ex))
        return {"fieldErrors": [{'name': field, 'status': 'Unknown error'}]}
